import os
import subprocess
import tempfile
from datetime import datetime, timezone
import re
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
import time
import urllib.parse
from urllib.parse import urlparse
import git
from utils.path_security import build_repository_local_path
from utils.security_utils import sanitize_text, sanitize_url

# 延迟导入pandas以避免版本冲突
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except Exception as e:
    from utils.safe_print import log_print
    log_print(f"警告: pandas导入失败: {e}", 'GIT', force=True)
    log_print("将使用openpyxl作为Excel处理的替代方案", 'GIT', force=True)
    PANDAS_AVAILABLE = False
    pd = None

# 配置GitPython以处理编码问题
git.cmd.Git.GIT_PYTHON_REFRESH = True
os.environ['PYTHONIOENCODING'] = 'utf-8'

# 修复GitPython的编码问题
import locale
import sys

# 设置系统编码
if sys.platform.startswith('win'):
    # Windows系统设置
    os.environ['PYTHONLEGACYWINDOWSSTDIO'] = '1'
    try:
        locale.setlocale(locale.LC_ALL, 'C.UTF-8')
    except:
        try:
            locale.setlocale(locale.LC_ALL, 'en_US.UTF-8')
        except:
            pass

# 猴子补丁GitPython的编码处理
original_execute = git.cmd.Git.execute

def patched_execute(self, command, *args, **kwargs):
    """修复编码问题的execute方法"""
    # 设置环境变量
    env = kwargs.get('env', os.environ.copy())
    env.update({
        'LC_ALL': 'C.UTF-8',
        'LANG': 'C.UTF-8',
        'PYTHONIOENCODING': 'utf-8'
    })
    kwargs['env'] = env
    
    try:
        return original_execute(self, command, *args, **kwargs)
    except UnicodeDecodeError as e:
        from utils.safe_print import log_print
        log_print(f"Git命令编码错误，尝试忽略: {e}", 'GIT', force=True)
        # 如果是编码错误，尝试用更宽松的方式执行
        if 'encoding' not in kwargs:
            kwargs['encoding'] = 'utf-8'
        if 'errors' not in kwargs:
            kwargs['errors'] = 'replace'
        return original_execute(self, command, *args, **kwargs)

# 应用补丁
git.cmd.Git.execute = patched_execute

class GitService:
    def __init__(self, repo_url, root_directory=None, username=None, token=None, repository=None, active_processes=None, max_workers=None):
        self.repo_url = repo_url
        self.root_directory = root_directory
        self.username = username
        self.token = token
        self.repository = repository
        self.active_processes = active_processes or set()
        
        # 解析仓库URL获取本地路径
        self.local_path = self._get_local_path()
        
        # 线程池配置（支持子类注入，避免父子类配置不一致）
        default_workers = min(32, (os.cpu_count() or 1) + 4)  # 基于CPU核心数设置线程数
        if max_workers is None:
            self.max_workers = default_workers
        else:
            self.max_workers = max(1, int(max_workers))
        # 延迟创建线程池，避免大量实例初始化时抢占线程资源
        self.thread_pool = None
        
        # 性能监控
        self.performance_stats = {
            'total_diff_time': 0,
            'excel_processing_time': 0,
            'git_operations_time': 0,
            'parallel_tasks_count': 0
        }

    def _get_thread_pool(self):
        """按需获取共享线程池，避免重复创建销毁开销。"""
        if self.thread_pool is None:
            self.thread_pool = ThreadPoolExecutor(max_workers=self.max_workers)
        return self.thread_pool
    
    def _get_local_path(self):
        if self.repository:
            # 兼容旧的初始化方式，并确保路径始终位于repos目录下
            return build_repository_local_path(
                self.repository.project.code,
                self.repository.name,
                self.repository.id,
                strict=False
            )
        else:
            # 新的初始化方式，用于获取分支信息
            return self.root_directory or 'temp_repo'
    
    def _run_git_command(self, cmd, cwd=None, timeout=300):
        """安全执行Git命令，处理编码问题"""
        try:
            from utils.safe_print import log_print
            log_print(f"🔧 执行Git命令: {sanitize_text(' '.join(cmd))}", 'GIT')
            log_print(f"🔧 工作目录: {cwd or self.local_path}", 'GIT')
            
            # 设置环境变量以处理中文编码
            env = os.environ.copy()
            env['LC_ALL'] = 'C.UTF-8'
            env['LANG'] = 'C.UTF-8'
            env['PYTHONIOENCODING'] = 'utf-8'
            
            result = subprocess.run(
                cmd,
                cwd=cwd or self.local_path,
                capture_output=True,
                text=True,
                encoding='utf-8',
                errors='replace',  # 替换无法解码的字符
                env=env,
                timeout=timeout
            )
            
            log_print(f"✅ Git命令完成，返回码: {result.returncode}", 'GIT')
            if result.stdout:
                log_print(f"📤 stdout: {sanitize_text(result.stdout[:200])}...", 'GIT')
            if result.stderr:
                # SSH 后量子加密警告等非错误信息不使用 force
                _stderr_is_warning = (result.returncode == 0 and
                    ('WARNING' in result.stderr or 'post-quantum' in result.stderr))
                log_print(f"📤 stderr: {sanitize_text(result.stderr[:200])}...",
                          'GIT', force=not _stderr_is_warning)
            
            return result
        except subprocess.TimeoutExpired:
            from utils.safe_print import log_print
            log_print(f"⏰ Git命令超时: {sanitize_text(' '.join(cmd))}", 'GIT', force=True)
            # 返回一个模拟的失败结果，而不是None
            class TimeoutResult:
                def __init__(self):
                    self.returncode = -1
                    self.stdout = ""
                    self.stderr = "命令执行超时"
            return TimeoutResult()
        except Exception as e:
            from utils.safe_print import log_print
            log_print(f"Git命令执行失败: {e}", 'GIT', force=True)
            return None
        
    def test_network_connectivity(self):
        """测试网络连接"""
        try:
            parsed_url = urlparse(self.repo_url)
            if parsed_url.scheme == 'ssh':
                hostname = parsed_url.hostname
                port = parsed_url.port or 22
                
                from utils.safe_print import log_print
                log_print(f"=== 网络连接诊断 ===", 'GIT')
                log_print(f"目标主机: {hostname}", 'GIT')
                log_print(f"端口: {port}", 'GIT')
                
                # 1. 测试DNS解析
                try:
                    import socket
                    ip = socket.gethostbyname(hostname)
                    log_print(f"DNS解析成功: {hostname} -> {ip}", 'GIT')
                    dns_ok = True
                except socket.gaierror as e:
                    log_print(f"DNS解析失败: {e}", 'GIT', force=True)
                    dns_ok = False
                
                # 2. 测试端口连通性
                if dns_ok:
                    try:
                        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                        sock.settimeout(10)
                        result = sock.connect_ex((hostname, port))
                        sock.close()
                        if result == 0:
                            log_print(f"端口连通性测试成功: {hostname}:{port}", 'GIT')
                            port_ok = True
                        else:
                            log_print(f"端口连通性测试失败: {hostname}:{port}", 'GIT', force=True)
                            port_ok = False
                    except Exception as e:
                        log_print(f"端口测试异常: {e}", 'GIT', force=True)
                        port_ok = False
                else:
                    port_ok = False
                
                return dns_ok and port_ok, "网络连接正常"
            else:
                log_print("非SSH URL，跳过网络连接测试", 'GIT')
                return True, "网络连接正常"
        except Exception as e:
            log_print(f"网络连接测试异常: {e}", 'GIT', force=True)
            return False, f"网络连接异常: {str(e)}"
    
    def get_branches(self):
        """获取远程仓库的分支列表"""
        try:
            # 使用git ls-remote命令获取远程分支
            cmd = ['git', 'ls-remote', '--heads', self.repo_url]
            
            # 如果有认证信息，构建带认证的URL
            if self.username and self.token:
                from urllib.parse import urlparse, urlunparse
                parsed = urlparse(self.repo_url)
                if parsed.scheme in ['http', 'https']:
                    # 构建带认证的URL
                    netloc = f"{self.username}:{self.token}@{parsed.netloc}"
                    auth_url = urlunparse((parsed.scheme, netloc, parsed.path, parsed.params, parsed.query, parsed.fragment))
                    cmd = ['git', 'ls-remote', '--heads', auth_url]
            
            # 设置环境变量和编码处理
            env = os.environ.copy()
            env['LC_ALL'] = 'C.UTF-8'
            env['LANG'] = 'C.UTF-8'
            env['PYTHONIOENCODING'] = 'utf-8'

            result = subprocess.run(
                cmd,
                capture_output=True,
                text=True,
                timeout=30,
                env=env,
                encoding='utf-8',
                errors='replace'  # 遇到无法解码的字符时用替换字符代替
            )
            
            if result.returncode == 0:
                branches = []
                for line in result.stdout.strip().split('\n'):
                    if line and 'refs/heads/' in line:
                        # 提取分支名称
                        branch_name = line.split('refs/heads/')[-1].strip()
                        if branch_name:
                            branches.append(branch_name)
                
                # 对分支进行排序，master和main优先
                def sort_branches(branch):
                    if branch == 'master':
                        return (0, branch)
                    elif branch == 'main':
                        return (1, branch)
                    else:
                        return (2, branch)
                
                branches.sort(key=sort_branches)
                return branches
            else:
                from utils.safe_print import log_print
                log_print(f"获取分支失败: {sanitize_text(result.stderr)}", 'GIT', force=True)
                return ['master']  # 返回默认分支
                
        except Exception as e:
            from utils.safe_print import log_print
            log_print(f"获取分支异常: {e}", 'GIT', force=True)
            return ['master']  # 返回默认分支

    def test_ssh_connection(self):
        """测试SSH连接"""
        try:
            repo_url = self.repo_url or (self.repository.url if self.repository else '')
            from utils.safe_print import log_print
            log_print(f"测试仓库URL: {sanitize_url(repo_url)}", 'GIT')
            
            # 检查是否为SSH格式的Git URL (git@hostname:path 或 ssh://git@hostname/path)
            if repo_url.startswith('git@') or repo_url.startswith('ssh://'):
                if repo_url.startswith('git@'):
                    # 处理 git@hostname:path 格式
                    parts = repo_url.split('@', 1)[1].split(':', 1)
                    hostname = parts[0]
                    username = 'git'
                    port = 22
                else:
                    # 处理 ssh://git@hostname/path 格式
                    parsed_url = urlparse(repo_url)
                    hostname = parsed_url.hostname
                    username = parsed_url.username or 'git'
                    port = parsed_url.port or 22
                
                log_print(f"检测到SSH URL，测试SSH连接到 {username}@{hostname}:{port}", 'GIT')
                
                # 使用git ls-remote测试连接，这更准确反映git操作的连接性
                try:
                    cmd = ['git', 'ls-remote', '--heads', repo_url]
                    # 设置环境变量和编码处理
                    env = os.environ.copy()
                    env['LC_ALL'] = 'C.UTF-8'
                    env['LANG'] = 'C.UTF-8'
                    env['PYTHONIOENCODING'] = 'utf-8'
                    
                    result = subprocess.run(
                        cmd, 
                        capture_output=True, 
                        text=True, 
                        timeout=30,
                        env=env,
                        encoding='utf-8',
                        errors='replace'  # 遇到无法解码的字符时用替换字符代替
                    )
                    
                    log_print(f"Git连接测试命令: {' '.join(cmd)}", 'GIT')
                    log_print(f"返回码: {result.returncode}", 'GIT')
                    if result.stdout:
                        log_print(f"输出: {sanitize_text(result.stdout[:200])}...", 'GIT')  # 只显示前200字符
                    if result.stderr:
                        # SSH 后量子加密警告在返回码=0时不当作错误
                        _is_ssh_warning = (result.returncode == 0 and
                            ('WARNING' in result.stderr or 'post-quantum' in result.stderr))
                        if _is_ssh_warning:
                            log_print(f"SSH连接警告(可忽略): {sanitize_text(result.stderr[:100])}...", 'GIT')
                        else:
                            log_print(f"错误: {sanitize_text(result.stderr)}", 'GIT', force=True)
                    
                    return result.returncode == 0
                except subprocess.TimeoutExpired:
                    log_print("Git连接测试超时", 'GIT', force=True)
                    return False
            else:
                log_print(f"非SSH URL ({sanitize_url(repo_url)})，跳过SSH连接测试", 'GIT')
                return True
        except Exception as e:
            log_print(f"SSH连接测试失败: {str(e)}", 'GIT', force=True)
            return False

    @staticmethod
    def _git_cmd_success(result):
        return bool(result and getattr(result, "returncode", -1) == 0)

    def _cleanup_git_lock_files(self):
        """清理常见的Git锁文件，防止pull/fetch被锁阻塞。"""
        removed_locks = []
        lock_files = [
            os.path.join(".git", "index.lock"),
            os.path.join(".git", "config.lock"),
            os.path.join(".git", "HEAD.lock"),
            os.path.join(".git", "packed-refs.lock"),
            os.path.join(".git", "shallow.lock"),
        ]
        for relative_path in lock_files:
            lock_path = os.path.join(self.local_path, relative_path)
            if not os.path.exists(lock_path):
                continue
            try:
                os.remove(lock_path)
                removed_locks.append(relative_path.replace("\\", "/"))
            except Exception:
                continue
        return removed_locks

    def _checkout_configured_branch(self):
        """显式切换到仓库配置的分支，避免在错误分支上pull。"""
        branch = (getattr(self.repository, "branch", "") or "").strip() if self.repository else ""
        if not branch:
            return True, "未配置分支，保持当前HEAD"

        self._run_git_command(['git', 'fetch', '--all', '--prune'], timeout=120)
        checkout_result = self._run_git_command(['git', 'checkout', branch], timeout=90)
        if self._git_cmd_success(checkout_result):
            return True, f"已切换分支: {branch}"

        recreate_result = self._run_git_command(['git', 'checkout', '-B', branch, f'origin/{branch}'], timeout=90)
        if self._git_cmd_success(recreate_result):
            return True, f"已重建并切换分支: {branch}"

        error_text = ""
        if recreate_result and getattr(recreate_result, "stderr", ""):
            error_text = recreate_result.stderr.strip()
        elif checkout_result and getattr(checkout_result, "stderr", ""):
            error_text = checkout_result.stderr.strip()
        return False, f"分支切换失败({branch}){': ' + sanitize_text(error_text) if error_text else ''}"

    def _pull_repository(self, timeout=120):
        branch = (getattr(self.repository, "branch", "") or "").strip() if self.repository else ""
        cmd = ['git', 'pull', '--no-rebase', 'origin']
        if branch:
            cmd.append(branch)
        return self._run_git_command(cmd, timeout=timeout)

    def _self_heal_repository_state(self):
        """失败后执行保守自愈，尽量恢复到可pull状态。"""
        from utils.safe_print import log_print

        removed_locks = self._cleanup_git_lock_files()
        if removed_locks:
            log_print(f"🧹 清理Git锁文件: {', '.join(removed_locks)}", 'GIT')

        reset_result = self._run_git_command(['git', 'reset', '--hard', 'HEAD'], timeout=90)
        clean_result = self._run_git_command(['git', 'clean', '-fd'], timeout=90)
        self._run_git_command(['git', 'gc', '--prune=now'], timeout=180)

        if not self._git_cmd_success(reset_result) or not self._git_cmd_success(clean_result):
            return False, "Git自愈失败(reset/clean未成功)"

        branch_ok, branch_message = self._checkout_configured_branch()
        if not branch_ok:
            return False, branch_message
        return True, "Git自愈完成"

    def clone_or_update_repository(self):
        """克隆或更新本地仓库"""
        try:
            # 避免循环导入，直接使用print
            print(f"🔧 [GIT_SERVICE] 进入 clone_or_update_repository 方法")
            print(f"🔧 [GIT_SERVICE] 本地路径: {self.local_path}")
            print(f"🔧 [GIT_SERVICE] 路径是否存在: {os.path.exists(self.local_path)}")
            
            # 使用utils中的log_print避免循环导入
            try:
                from utils.safe_print import log_print
                log_print(f"检查本地路径: {self.local_path}", 'GIT')
                log_print(f"路径是否存在: {os.path.exists(self.local_path)}", 'GIT')
            except ImportError:
                print(f"🔧 [GIT_SERVICE] 检查本地路径: {self.local_path}")
                print(f"🔧 [GIT_SERVICE] 路径是否存在: {os.path.exists(self.local_path)}")
            
            if os.path.exists(self.local_path):
                # 如果本地仓库已存在，则更新
                log_print("本地仓库已存在，开始更新...", 'GIT')

                removed_locks = self._cleanup_git_lock_files()
                if removed_locks:
                    log_print(f"🧹 检测并清理Git锁文件: {', '.join(removed_locks)}", 'GIT')

                branch_ok, branch_message = self._checkout_configured_branch()
                if not branch_ok:
                    log_print(branch_message, 'GIT', force=True)
                    return False, branch_message

                # 使用显式分支pull，避免拉错分支
                result = self._pull_repository(timeout=120)

                if self._git_cmd_success(result):
                    log_print(f"仓库已更新: {self.local_path}", 'GIT')
                    return True, "仓库更新成功"
                elif result:
                    log_print(f"Git pull输出: {result.stdout}", 'GIT')
                    log_print(f"Git pull错误: {result.stderr}", 'GIT')

                    # 超时视为失败，进入自愈流程
                    if result.returncode == -1 and "超时" in result.stderr:
                        log_print("Git命令超时，尝试执行仓库自愈", 'GIT', force=True)

                    # 检查是否实际更新成功（有时stderr有内容但实际成功了）
                    if "Already up to date" in result.stdout or "Already up-to-date" in result.stdout:
                        log_print("仓库已是最新状态", 'GIT')
                        return True, "仓库已是最新状态"
                    elif "Fast-forward" in result.stdout or result.returncode == 0:
                        log_print("仓库更新成功", 'GIT')
                        return True, "仓库更新成功"
                    else:
                        # pull失败后，执行一次自愈并重试
                        log_print("尝试自愈仓库状态并重试更新...", 'GIT')
                        try:
                            heal_ok, heal_msg = self._self_heal_repository_state()
                            if not heal_ok:
                                return False, f"仓库自愈失败: {heal_msg}"

                            pull_result = self._pull_repository(timeout=120)
                            if self._git_cmd_success(pull_result):
                                log_print(f"仓库已更新: {self.local_path}", 'GIT')
                                return True, "仓库更新成功（自愈后）"

                            # 兜底尝试GitPython
                            repo = git.Repo(self.local_path)
                            origin = repo.remotes.origin
                            branch = (getattr(self.repository, "branch", "") or "").strip() if self.repository else ""
                            if branch:
                                origin.pull(branch)
                            else:
                                origin.pull()
                            log_print(f"仓库已更新: {self.local_path}", 'GIT')
                            return True, "仓库更新成功（GitPython兜底）"
                        except Exception as git_e:
                            log_print(f"GitPython更新失败: {git_e}", 'GIT')
                            # 如果是编码错误，忽略并认为更新成功
                            if any(keyword in str(git_e).lower() for keyword in ['unicodedecodeerror', 'gbk', 'encoding']):
                                log_print("检测到编码问题，但仓库可能已更新", 'GIT')
                                return True, "仓库更新成功（忽略编码警告）"
                            return False, f"仓库更新失败: {str(git_e)}"
                else:
                    print("Git命令执行失败")
                    return False, "Git命令执行失败"
            else:
                # 克隆新仓库
                print("本地仓库不存在，开始克隆...")
                print(f"创建目录: {os.path.dirname(self.local_path)}")
                os.makedirs(os.path.dirname(self.local_path), exist_ok=True)
                
                # 构建带认证的URL
                clone_url = self.repo_url or (self.repository.url if self.repository else '')
                
                # 检查URL类型并处理认证
                parsed_url = urllib.parse.urlparse(clone_url)
                print(f"原始URL: {sanitize_url(clone_url)}")
                print(f"URL scheme: {parsed_url.scheme}")
                print(f"URL host: {parsed_url.hostname or parsed_url.netloc}")
                
                if self.repository.token and parsed_url.scheme in ['http', 'https']:
                    # 对于HTTP/HTTPS URL，使用token认证
                    if parsed_url.scheme and parsed_url.netloc:
                        clone_url = f"{parsed_url.scheme}://oauth2:{self.repository.token}@{parsed_url.netloc}{parsed_url.path}"
                        print(f"使用HTTP token认证")
                    else:
                        print(f"HTTP URL格式可能有问题: {sanitize_url(self.repo_url or '')}")
                elif parsed_url.scheme == 'ssh':
                    # 对于SSH URL，直接使用原URL，依赖系统SSH配置
                    print(f"使用SSH认证，依赖系统SSH密钥配置")
                    clone_url = self.repo_url
                else:
                    print(f"使用原始URL，无特殊认证处理")
                
                print(f"最终克隆URL: {sanitize_url(clone_url)}")
                print(f"目标路径: {self.local_path}")
                print(f"分支: {self.repository.branch}")
                
                # 尝试克隆
                try:
                    if self.repository.branch:
                        git.Repo.clone_from(clone_url, self.local_path, branch=self.repository.branch)
                    else:
                        git.Repo.clone_from(clone_url, self.local_path)
                except git.exc.GitCommandError as git_error:
                    # 如果是SSH连接问题，提供更详细的错误信息
                    if 'Could not resolve hostname' in str(git_error) or 'Name or service not known' in str(git_error):
                        raise Exception(f"SSH主机名解析失败，请检查：\n1. 网络连接是否正常\n2. SSH配置是否正确\n3. 主机名是否可访问\n原始错误: {sanitize_text(str(git_error))}")
                    elif 'Permission denied' in str(git_error) or 'access rights' in str(git_error):
                        raise Exception(f"SSH认证失败，请检查：\n1. SSH密钥是否正确配置\n2. 是否有仓库访问权限\n原始错误: {sanitize_text(str(git_error))}")
                    else:
                        raise git_error
                
                print(f"仓库已克隆: {self.local_path}")
                return True, "仓库克隆成功"
            
        except git.exc.GitCommandError as e:
            error_msg = f"Git命令错误: {sanitize_text(str(e))}"
            print(error_msg)
            return False, error_msg
        except Exception as e:
            error_msg = f"仓库同步失败: {sanitize_text(str(e))}"
            print(error_msg)
            import traceback
            traceback.print_exc()
            return False, error_msg
    
    def get_commits(self, since_date=None, limit=100):
        """获取提交记录"""
        try:
            if not os.path.exists(self.local_path):
                success, message = self.clone_or_update_repository()
                if not success:
                    return []
            
            # 配置Git仓库以处理编码
            repo = git.Repo(self.local_path)
            # 设置Git配置以处理中文编码
            try:
                repo.git.config('core.quotepath', 'false')
                repo.git.config('i18n.commitencoding', 'utf-8')
                repo.git.config('i18n.logoutputencoding', 'utf-8')
            except:
                pass  # 忽略配置错误
            
            commits = []
            
            # 获取指定分支的提交记录
            branch = repo.heads[self.repository.branch] if self.repository.branch in [h.name for h in repo.heads] else repo.head
            
            # 构建iter_commits参数
            iter_kwargs = {'max_count': limit}
            if since_date:
                iter_kwargs['since'] = since_date
                print(f"🔍 [GIT_SERVICE] 增量同步，从 {since_date} 开始获取最多 {limit} 个提交")
            else:
                print(f"🔍 [GIT_SERVICE] 全量同步，获取最多 {limit} 个提交")
            
            commit_iter = repo.iter_commits(branch, **iter_kwargs)
            
            processed_count = 0
            for commit in commit_iter:
                # 确保不超过limit限制
                if processed_count >= limit:
                    print(f"🔍 [GIT_SERVICE] 已达到限制数量 {limit}，停止获取")
                    break
                processed_count += 1
                try:
                    # 验证提交对象是否有效
                    _ = commit.hexsha
                    _ = commit.author.name
                    _ = commit.message
                except Exception as commit_error:
                    print(f"⚠️ [GIT_SERVICE] 跳过无效提交: {commit_error}")
                    continue
                
                # 过滤提交人
                if self.repository.commit_filter:
                    filter_emails = [email.strip() for email in self.repository.commit_filter.split(',')]
                    if commit.author.email in filter_emails:
                        continue
                
                # 过滤日志
                if self.repository.log_filter_regex:
                    import re
                    if re.match(self.repository.log_filter_regex, commit.message):
                        continue
                
                # 获取文件变更 - 使用diff来正确检测操作类型
                if commit.parents:
                    # 有父提交，比较差异
                    try:
                        parent = commit.parents[0]
                        # 验证父提交是否有效
                        try:
                            _ = parent.hexsha
                            _ = parent.tree
                        except Exception as parent_error:
                            print(f"⚠️ [GIT_SERVICE] 提交 {commit.hexsha[:8]} 的父提交无效: {parent_error}")
                            raise Exception(f"父提交无效: {parent_error}")
                        
                        diffs = parent.diff(commit)
                    except Exception as diff_error:
                        print(f"⚠️ [GIT_SERVICE] 提交 {commit.hexsha[:8]} diff比较失败: {diff_error}")
                        # 使用stats作为备选方案
                        try:
                            for file_path in commit.stats.files:
                                if self.repository.path_regex:
                                    import re
                                    if not re.search(self.repository.path_regex, file_path):
                                        continue
                                
                                commits.append({
                                    'commit_id': commit.hexsha,
                                    'path': file_path,
                                    'version': commit.hexsha[:8],
                                    'operation': 'M',  # 默认为修改
                                    'author': commit.author.name,
                                    'author_email': commit.author.email,
                                    'commit_time': datetime.fromtimestamp(commit.committed_date, tz=timezone.utc),
                                    'message': commit.message.strip()
                                })
                        except Exception as stats_error:
                            print(f"⚠️ [GIT_SERVICE] 提交 {commit.hexsha[:8]} stats获取也失败: {stats_error}")
                            # 如果stats也失败，至少记录这个提交存在
                            print(f"⚠️ [GIT_SERVICE] 将提交 {commit.hexsha[:8]} 标记为存在但无法分析文件变更")
                        continue
                    
                    for diff in diffs:
                        file_path = diff.b_path or diff.a_path
                        
                        # 路径过滤
                        if self.repository.path_regex:
                            import re
                            if not re.search(self.repository.path_regex, file_path):
                                continue
                        
                        # 根据diff类型确定操作
                        if diff.change_type == 'A':
                            operation = 'A'  # 新增
                        elif diff.change_type == 'D':
                            operation = 'D'  # 删除
                        elif diff.change_type == 'M':
                            operation = 'M'  # 修改
                        elif diff.change_type == 'R':
                            operation = 'M'  # 重命名视为修改
                        else:
                            operation = 'M'  # 其他情况默认为修改
                        
                        commits.append({
                            'commit_id': commit.hexsha,
                            'path': file_path,
                            'version': commit.hexsha[:8],
                            'operation': operation,
                            'author': commit.author.name,
                            'author_email': commit.author.email,
                            'commit_time': datetime.fromtimestamp(commit.committed_date, tz=timezone.utc),
                            'message': commit.message.strip()
                        })
                else:
                    # 初始提交，所有文件都是新增
                    for item in commit.stats.files:
                        file_path = item
                        
                        # 路径过滤
                        if self.repository.path_regex:
                            import re
                            if not re.search(self.repository.path_regex, file_path):
                                continue
                        
                        operation = 'A'  # 初始提交都是新增
                        
                        commits.append({
                            'commit_id': commit.hexsha,
                            'path': file_path,
                            'version': commit.hexsha[:8],
                            'operation': operation,
                            'author': commit.author.name,
                            'author_email': commit.author.email,
                            'commit_time': datetime.fromtimestamp(commit.committed_date, tz=timezone.utc),
                            'message': commit.message.strip()
                        })
            
            # 为每个文件收集前一次提交记录
            commits = self._collect_previous_commits(repo, commits)
            
            return commits
        except Exception as e:
            print(f"获取提交记录失败: {str(e)}")
            return []
    
    def _collect_previous_commits(self, repo, commits):
        """为每个文件收集前一次的提交记录，确保能找到对比版本"""
        try:
            print(f"🔍 [GIT_SERVICE] 开始收集文件的前一次提交记录...")
            
            # 按文件路径分组
            files_commits = {}
            for commit_data in commits:
                file_path = commit_data['path']
                if file_path not in files_commits:
                    files_commits[file_path] = []
                files_commits[file_path].append(commit_data)
            
            additional_commits = []
            
            for file_path, file_commits in files_commits.items():
                # 按时间排序，最新的在前
                file_commits.sort(key=lambda x: x['commit_time'], reverse=True)
                
                # 获取最早的提交
                earliest_commit = file_commits[-1]
                earliest_commit_obj = repo.commit(earliest_commit['commit_id'])
                
                # 查找这个文件的前一次提交
                try:
                    # 使用git log --follow来跟踪文件历史，包括重命名
                    git_cmd = repo.git
                    log_output = git_cmd.log(
                        '--follow', 
                        '--format=%H|%an|%ae|%ct|%s',
                        '--', file_path
                    )
                    
                    if log_output:
                        log_lines = log_output.strip().split('\n')
                        
                        # 找到当前最早提交在历史中的位置
                        earliest_commit_index = -1
                        for i, line in enumerate(log_lines):
                            if line.startswith(earliest_commit['commit_id']):
                                earliest_commit_index = i
                                break
                        
                        # 如果找到了，并且还有更早的提交
                        if earliest_commit_index >= 0 and earliest_commit_index < len(log_lines) - 1:
                            prev_line = log_lines[earliest_commit_index + 1]
                            parts = prev_line.split('|')
                            if len(parts) >= 5:
                                prev_commit_id = parts[0]
                                prev_author = parts[1]
                                prev_author_email = parts[2]
                                prev_timestamp = int(parts[3])
                                prev_message = '|'.join(parts[4:])
                                
                                # 检查这个前一次提交是否已经在结果中
                                already_exists = any(
                                    c['commit_id'] == prev_commit_id and c['path'] == file_path 
                                    for c in commits + additional_commits
                                )
                                
                                if not already_exists:
                                    # 确定操作类型
                                    prev_commit_obj = repo.commit(prev_commit_id)
                                    operation = 'M'  # 默认为修改
                                    
                                    # 如果这是文件的第一次提交（没有父提交），则为新增
                                    if not prev_commit_obj.parents:
                                        operation = 'A'
                                    
                                    additional_commits.append({
                                        'commit_id': prev_commit_id,
                                        'path': file_path,
                                        'version': prev_commit_id[:8],
                                        'operation': operation,
                                        'author': prev_author,
                                        'author_email': prev_author_email,
                                        'commit_time': datetime.fromtimestamp(prev_timestamp, tz=timezone.utc),
                                        'message': prev_message.strip()
                                    })
                                    
                                    print(f"📝 [GIT_SERVICE] 为文件 {file_path} 找到前一次提交: {prev_commit_id[:8]}")
                
                except Exception as e:
                    print(f"⚠️ [GIT_SERVICE] 获取文件 {file_path} 的前一次提交失败: {e}")
                    continue
            
            if additional_commits:
                print(f"🔍 [GIT_SERVICE] 额外收集了 {len(additional_commits)} 个前一次提交记录")
                commits.extend(additional_commits)
            
            return commits
            
        except Exception as e:
            print(f"⚠️ [GIT_SERVICE] 收集前一次提交记录失败: {e}")
            return commits
    
    def get_file_diff(self, commit_id, file_path):
        """获取文件的差异内容"""
        try:
            start_time = time.time()
            print(f"开始获取diff: commit_id={commit_id}, file_path={file_path}")
            
            if not os.path.exists(self.local_path):
                print(f"本地路径不存在: {self.local_path}")
                return None
            
            repo = git.Repo(self.local_path)
            commit = repo.commit(commit_id)
            print(f"找到提交: {commit.hexsha[:8]} - {commit.message.strip()}")
            
            # 获取当前提交的文件内容
            try:
                current_blob = commit.tree[file_path]
                current_content = current_blob.data_stream.read().decode('utf-8')
            except KeyError:
                print(f"文件在当前提交中不存在: {file_path}")
                current_content = ""
            except UnicodeDecodeError:
                print(f"文件编码问题，使用替换模式: {file_path}")
                current_blob = commit.tree[file_path]
                current_content = current_blob.data_stream.read().decode('utf-8', errors='replace')
            
            # 获取文件的diff
            if commit.parents:
                parent_commit = commit.parents[0]
                print(f"父提交: {parent_commit.hexsha[:8]}")
                
                # 获取父提交的文件内容
                try:
                    parent_blob = parent_commit.tree[file_path]
                    previous_content = parent_blob.data_stream.read().decode('utf-8')
                except KeyError:
                    print(f"文件在父提交中不存在，视为新增文件: {file_path}")
                    previous_content = ""
                except UnicodeDecodeError:
                    print(f"父提交文件编码问题，使用替换模式: {file_path}")
                    parent_blob = parent_commit.tree[file_path]
                    previous_content = parent_blob.data_stream.read().decode('utf-8', errors='replace')
                
                # 使用GitPython生成diff
                try:
                    diffs = parent_commit.diff(commit, paths=[file_path], create_patch=True)
                    print(f"找到 {len(diffs)} 个diff")
                    
                    if diffs:
                        diff = diffs[0]
                        
                        # 获取patch内容
                        if hasattr(diff, 'diff') and diff.diff:
                            patch_text = diff.diff.decode('utf-8') if isinstance(diff.diff, bytes) else str(diff.diff)
                        else:
                            # 备用方案：使用subprocess生成diff
                            try:
                                cmd = ['git', 'diff', parent_commit.hexsha, commit.hexsha, '--', file_path]
                                result = self._run_git_command(cmd, timeout=30)
                                if result and result.returncode == 0:
                                    patch_text = result.stdout
                                else:
                                    print(f"Git diff命令失败: {result.stderr if result else 'Command failed'}")
                                    return None
                            except Exception as e:
                                print(f"执行Git diff命令失败: {str(e)}")
                                return None
                        
                        # 解析patch内容
                        if patch_text:
                            hunks = self._parse_unified_diff(patch_text)
                            diff_result = {
                                'type': 'code',
                                'file_path': file_path,
                                'patch': patch_text,
                                'hunks': hunks
                            }
                        else:
                            # 如果没有patch内容，生成基本的diff结构
                            diff_result = self._generate_basic_diff(previous_content, current_content, file_path)
                    else:
                        print("未找到diff，可能文件无变化")
                        return None
                        
                except Exception as git_e:
                    print(f"GitPython diff失败: {git_e}")
                    # 使用基本diff生成作为备用
                    diff_result = self._generate_basic_diff(previous_content, current_content, file_path)
            else:
                # 初始提交，所有内容都是新增的
                print("初始提交，生成新增文件的diff")
                diff_result = self._generate_initial_commit_diff(current_content, file_path)
            
            # 更新性能统计
            processing_time = time.time() - start_time
            self.performance_stats['git_operations_time'] += processing_time
            self.performance_stats['total_diff_time'] += processing_time
            
            return diff_result
            
        except Exception as e:
            print(f"获取文件diff失败: {str(e)}")
            import traceback
            traceback.print_exc()
            return None
    
    def get_parent_commit(self, commit_id):
        """获取指定提交的父提交ID"""
        try:
            if not os.path.exists(self.local_path):
                return None
                
            repo = git.Repo(self.local_path)
            commit = repo.commit(commit_id)
            
            if commit.parents:
                return commit.parents[0].hexsha
            return None
            
        except Exception as e:
            return None

    @staticmethod
    def _commit_id_matches(candidate_commit_id, target_commit_id):
        """判断两个 commit id 是否匹配（支持长短 SHA 前缀匹配）。"""
        candidate = str(candidate_commit_id or "").strip().lower()
        target = str(target_commit_id or "").strip().lower()
        if not candidate or not target:
            return False
        return candidate == target or candidate.startswith(target) or target.startswith(candidate)

    def get_previous_file_commit(self, file_path, current_commit_id, max_count=5000):
        """按文件历史查找当前提交的上一提交（不受业务 start_date 限制）。"""
        try:
            if not file_path or not current_commit_id:
                return None

            if not os.path.exists(self.local_path):
                clone_ok, _ = self.clone_or_update_repository()
                if not clone_ok:
                    return None

            git_cmd = [
                "git",
                "log",
                "--follow",
                f"--max-count={max(50, int(max_count or 5000))}",
                "--format=%H|%an|%ct|%s",
                str(current_commit_id),
                "--",
                str(file_path),
            ]
            result = self._run_git_command(git_cmd, timeout=180)
            if not result or result.returncode != 0 or not result.stdout:
                return None

            log_lines = [line.strip() for line in str(result.stdout).splitlines() if line.strip()]
            current_index = None
            for idx, line in enumerate(log_lines):
                commit_id = line.split("|", 1)[0].strip()
                if self._commit_id_matches(commit_id, current_commit_id):
                    current_index = idx
                    break

            if current_index is None or current_index + 1 >= len(log_lines):
                return None

            previous_line = log_lines[current_index + 1]
            parts = previous_line.split("|")
            if len(parts) < 4:
                return None

            previous_commit_id = parts[0].strip()
            author = parts[1].strip()
            timestamp_text = parts[2].strip()
            message = "|".join(parts[3:]).strip() if len(parts) > 3 else ""

            commit_time = None
            try:
                commit_time = datetime.fromtimestamp(int(timestamp_text), tz=timezone.utc)
            except Exception:
                commit_time = None

            return {
                "commit_id": previous_commit_id,
                "author": author,
                "message": message,
                "commit_time": commit_time,
            }
        except Exception as e:
            from utils.safe_print import log_print
            log_print(f"获取文件上一提交失败: {e}", "GIT")
            return None
    
    def get_commit_range_diff(self, from_commit, to_commit, file_path):
        """获取提交范围内的diff"""
        try:
            print(f"=== Git diff range调试信息 ===")
            print(f"本地路径: {self.local_path}")
            print(f"路径存在: {os.path.exists(self.local_path)}")
            print(f"从提交: {from_commit}")
            print(f"到提交: {to_commit}")
            print(f"文件路径: {file_path}")
            
            if not os.path.exists(self.local_path):
                print("✗ 本地路径不存在")
                return None
                
            repo = git.Repo(self.local_path)
            print(f"Git仓库初始化成功")
            
            # 验证提交是否存在
            try:
                from_commit_obj = repo.commit(from_commit)
                to_commit_obj = repo.commit(to_commit)
                print(f"✓ 提交验证成功:")
                print(f"  - 从提交: {from_commit_obj.hexsha[:8]} - {from_commit_obj.message.strip()[:50]}")
                print(f"  - 到提交: {to_commit_obj.hexsha[:8]} - {to_commit_obj.message.strip()[:50]}")
            except Exception as e:
                print(f"✗ 提交验证失败: {e}")
                print(f"尝试查找类似的提交...")
                
                # 尝试查找类似的提交（前8位匹配）
                try:
                    all_commits = list(repo.iter_commits('--all', max_count=200))
                    print(f"仓库中共有 {len(all_commits)} 个提交")
                    
                    from_short = from_commit[:8]
                    to_short = to_commit[:8]
                    
                    found_from = None
                    found_to = None
                    
                    for commit in all_commits:
                        if commit.hexsha.startswith(from_short):
                            found_from = commit.hexsha
                            print(f"找到匹配的from提交: {found_from[:8]} - {commit.message.strip()[:50]}")
                        if commit.hexsha.startswith(to_short):
                            found_to = commit.hexsha
                            print(f"找到匹配的to提交: {found_to[:8]} - {commit.message.strip()[:50]}")
                    
                    if found_from and found_to:
                        print(f"使用找到的提交进行diff")
                        from_commit = found_from
                        to_commit = found_to
                    else:
                        print(f"✗ 未找到匹配的提交")
                        return None
                        
                except Exception as search_e:
                    print(f"✗ 搜索提交失败: {search_e}")
                    return None
            
            # 使用GitPython直接执行diff，避免subprocess环境问题
            try:
                import time
                start_time = time.time()
                
                print(f"使用GitPython执行diff: {from_commit[:8]}..{to_commit[:8]} -- {file_path}")
                
                # 直接使用GitPython的diff功能
                from_commit_obj = repo.commit(from_commit)
                to_commit_obj = repo.commit(to_commit)
                
                # 获取两个提交之间的diff
                diffs = from_commit_obj.diff(to_commit_obj, paths=[file_path], create_patch=True)
                
                if diffs:
                    diff = diffs[0]
                    
                    # 获取patch内容
                    if hasattr(diff, 'diff') and diff.diff:
                        diff_output = diff.diff.decode('utf-8') if isinstance(diff.diff, bytes) else str(diff.diff)
                    else:
                        # 使用repo.git.diff作为备用
                        diff_output = repo.git.diff(from_commit, to_commit, file_path, unified=3)
                    
                    end_time = time.time()
                    print(f"✓ GitPython diff执行成功，耗时: {end_time - start_time:.2f}秒")
                    print(f"输出长度: {len(diff_output)} 字符")
                    
                else:
                    print("✗ 未找到diff内容")
                    return None
                    
            except Exception as git_e:
                end_time = time.time()
                print(f"✗ GitPython diff失败，耗时: {end_time - start_time:.2f}秒，错误: {git_e}")
                
                # 备用方案：使用repo.git.diff
                try:
                    print("尝试备用方案：使用repo.git.diff")
                    start_time = time.time()
                    diff_output = repo.git.diff(from_commit, to_commit, file_path, unified=3)
                    end_time = time.time()
                    print(f"✓ 备用方案成功，耗时: {end_time - start_time:.2f}秒")
                    print(f"输出长度: {len(diff_output)} 字符")
                except Exception as backup_e:
                    print(f"✗ 备用方案也失败: {backup_e}")
                    return None
            
            if not diff_output:
                return None
                
            # 解析diff输出
            hunks = self._parse_unified_diff(diff_output)
            
            diff_data = {
                'type': 'code',
                'file_path': file_path,
                'patch': diff_output,
                'hunks': hunks
            }
            
            return diff_data
            
        except Exception as e:
            return None

    def get_commit_info(self, commit_id):
        """获取单个commit的详细信息"""
        try:
            if not os.path.exists(self.local_path):
                print(f"✗ 本地路径不存在: {self.local_path}")
                return None

            repo = git.Repo(self.local_path)

            # 验证并获取commit对象
            try:
                commit_obj = repo.commit(commit_id)
            except Exception as e:
                print(f"✗ 提交验证失败: {e}")
                # 尝试查找类似的提交（前8位匹配）
                try:
                    all_commits = list(repo.iter_commits('--all', max_count=200))
                    commit_short = commit_id[:8]

                    for commit in all_commits:
                        if commit.hexsha.startswith(commit_short):
                            commit_obj = commit
                            print(f"找到匹配的提交: {commit_obj.hexsha[:8]} - {commit_obj.message.strip()[:50]}")
                            break
                    else:
                        print(f"✗ 未找到匹配的提交")
                        return None

                except Exception as search_e:
                    print(f"✗ 搜索提交失败: {search_e}")
                    return None

            # 提取commit信息
            commit_info = {
                'commit_id': commit_obj.hexsha,
                'short_id': commit_obj.hexsha[:8],
                'author': commit_obj.author.name,
                'author_email': commit_obj.author.email,
                'message': commit_obj.message.strip(),
                'commit_time': commit_obj.committed_datetime.strftime('%Y-%m-%d %H:%M:%S'),
                'commit_timestamp': commit_obj.committed_datetime
            }

            print(f"✓ 获取commit信息成功: {commit_info['short_id']} - {commit_info['author']} - {commit_info['message'][:50]}")
            return commit_info

        except Exception as e:
            print(f"✗ 获取commit信息失败: {e}")
            return None
    
    def get_file_content(self, commit_id, file_path):
        """获取指定提交中文件的内容"""
        try:
            if not os.path.exists(self.local_path):
                return None
            
            repo = git.Repo(self.local_path)
            commit = repo.commit(commit_id)
            
            try:
                blob = commit.tree[file_path]
                return blob.data_stream.read().decode('utf-8')
            except KeyError:
                return None
        except Exception as e:
            print(f"获取文件内容失败: {str(e)}")
            return None
    
    def parse_excel_diff(self, commit_id, file_path):
        """解析Excel文件的差异，返回表格对比数据"""
        try:
            print(f"开始解析Excel差异: commit_id={commit_id}, file_path={file_path}")
            if not os.path.exists(self.local_path):
                print(f"本地路径不存在: {self.local_path}")
                return None
            
            print("正在初始化Git仓库...")
            repo = git.Repo(self.local_path)
            
            try:
                print(f"正在获取commit: {commit_id}")
                commit = repo.commit(commit_id)
                
                # 检查文件是否在当前提交中存在（可能是删除操作）
                try:
                    commit.tree[file_path]
                    file_exists_in_current = True
                except KeyError:
                    file_exists_in_current = False
                    print(f"文件在当前提交中不存在，可能是删除操作: {file_path}")
                
                # 如果文件在当前提交中不存在，这是删除操作
                if not file_exists_in_current:
                    return {
                        'type': 'excel',
                        'file_type': 'Excel',
                        'operation': 'deleted',
                        'message': '该Excel文件已被删除',
                        'file_path': file_path
                    }
                
                print("正在提取当前版本Excel数据...")
                current_data = self._extract_excel_data(commit, file_path)
                
                previous_data = None
                if commit.parents:
                    parent_commit = commit.parents[0]
                    try:
                        previous_data = self._extract_excel_data(parent_commit, file_path)
                    except KeyError:
                        # 文件在父提交中不存在，是新增文件
                        pass
                
                # 生成表格对比数据
                return self._generate_excel_diff_data(current_data, previous_data, file_path)
                    
            except Exception as e:
                return {
                    'type': 'excel',
                    'file_type': 'Excel',
                    'error': True,
                    'message': f'无法解析Excel差异: {str(e)}'
                }
                
        except Exception as e:
            print(f"解析Excel差异失败: {str(e)}")
            return None
    
    def _extract_excel_data(self, commit, file_path):
        """从指定提交中提取Excel数据"""
        try:
            # 检查文件是否存在于该提交中
            try:
                blob = commit.tree / file_path
            except KeyError:
                print(f"文件在提交 {commit.hexsha[:8]} 中不存在: {file_path}")
                return None
                
            excel_data = blob.data_stream.read()
            
            # 使用openpyxl直接读取Excel数据，避免pandas/numpy兼容性问题
            import io
            import warnings
            # 抑制openpyxl的Data Validation警告
            warnings.filterwarnings('ignore', message='Data Validation extension is not supported and will be removed')
            from openpyxl import load_workbook
            
            excel_file = io.BytesIO(excel_data)
            
            # 读取所有工作表
            sheets_data = {}
            try:
                workbook = load_workbook(excel_file, data_only=True)
                
                for sheet_name in workbook.sheetnames:
                    worksheet = workbook[sheet_name]
                    
                    # 获取工作表数据 - 修复空数据问题
                    sheet_rows = []
                    max_row = worksheet.max_row
                    max_col = worksheet.max_column
                    
                    print(f"工作表 {sheet_name}: max_row={max_row}, max_col={max_col}")
                    
                    # 检查工作表是否真的有数据
                    if max_row > 0 and max_col > 0:
                        # 先检查是否有实际内容
                        has_content = False
                        for row in range(1, min(max_row + 1, 11)):  # 检查前10行
                            for col in range(1, min(max_col + 1, 11)):  # 检查前10列
                                cell_value = worksheet.cell(row=row, column=col).value
                                if cell_value is not None and str(cell_value).strip():
                                    has_content = True
                                    break
                            if has_content:
                                break
                        
                        if has_content:
                            # 智能检测实际数据边界
                            actual_bounds = self._detect_data_bounds(worksheet, max_row, max_col)
                            actual_max_row = actual_bounds['max_row']
                            actual_max_col = actual_bounds['max_col']
                            if 'rows' in sheets_data:
                                print(f"- 工作表行数: {len(sheets_data['rows'])}")
                                if sheets_data['rows']:
                                    print(f"- 第一行数据示例: {sheets_data['rows'][0]}")
                                else:
                                    print(f"- 工作表rows为空，状态: {sheets_data.get('status')}")
                                    print(f"- has_changes: {sheets_data.get('has_changes')}")
                            print(f"工作表 {sheet_name} 数据边界: 行1-{actual_max_row}, 列1-{actual_max_col}")
                            
                            for row in range(1, actual_max_row + 1):
                                row_data = {}
                                row_has_data = False
                                
                                for col in range(1, actual_max_col + 1):
                                    cell_value = worksheet.cell(row=row, column=col).value
                                    col_letter = self._get_column_letter(col)
                                    cell_str = str(cell_value) if cell_value is not None else ''
                                    row_data[col_letter] = cell_str
                                    if cell_str.strip():
                                        row_has_data = True
                                
                                # 添加所有行（包括空行），但记录数据边界
                                sheet_rows.append(row_data)
                        
                        print(f"工作表 {sheet_name} 实际读取到 {len(sheet_rows)} 行数据")
                    
                    sheets_data[sheet_name] = sheet_rows
                
                workbook.close()
                return sheets_data
                
            except Exception as e:
                print(f"使用openpyxl读取Excel失败: {str(e)}")
                # 如果openpyxl失败，尝试简化的处理方式
                return self._extract_excel_data_simple(excel_data, file_path)
            
        except Exception as e:
            print(f"提取Excel数据失败: {str(e)}")
            return None
    
    def _get_column_letter(self, col_num):
        """将列号转换为Excel列字母 (1->A, 2->B, ...)"""
        result = ""
        while col_num > 0:
            col_num -= 1
            result = chr(col_num % 26 + ord('A')) + result
            col_num //= 26
        return result
    
    def _detect_data_bounds(self, worksheet, max_row, max_col):
        """检测工作表的实际数据边界，排除空白行列"""
        actual_max_row = 0
        actual_max_col = 0
        
        # 从后往前扫描，找到最后一个有数据的行
        for row in range(max_row, 0, -1):
            has_data = False
            for col in range(1, max_col + 1):
                cell_value = worksheet.cell(row=row, column=col).value
                if cell_value is not None and str(cell_value).strip():
                    has_data = True
                    break
            if has_data:
                actual_max_row = row
                break
        
        # 从右往左扫描，找到最后一个有数据的列
        for col in range(max_col, 0, -1):
            has_data = False
            for row in range(1, actual_max_row + 1):
                cell_value = worksheet.cell(row=row, column=col).value
                if cell_value is not None and str(cell_value).strip():
                    has_data = True
                    break
            if has_data:
                actual_max_col = col
                break
        
        # 安全边界检查，至少保留一些列
        actual_max_col = max(actual_max_col, 10)
        actual_max_row = max(actual_max_row, 1)
        
        return {
            'max_row': actual_max_row,
            'max_col': actual_max_col
        }
    
    def _extract_excel_data_simple(self, excel_data, file_path):
        """简化的Excel数据提取，返回基本信息"""
        try:
            return {
                'Sheet1': [{
                    'A': f'Excel文件: {file_path}',
                    'B': f'文件大小: {len(excel_data)} 字节',
                    'C': '由于兼容性问题，无法显示详细内容'
                }]
            }
        except Exception as e:
            print(f"简化Excel数据提取失败: {str(e)}")
            return None
    
    def _generate_excel_diff_data(self, current_data, previous_data, file_path):
        """生成Excel差异对比数据 - 智能空白区域处理版本"""
        start_time = time.time()
        
        if not current_data:
            return {
                'type': 'excel',
                'file_type': 'Excel',
                'error': True,
                'message': '无法读取Excel文件内容'
            }
        
        if not previous_data:
            # 新增文件 - 优化显示边界
            optimized_sheets = self._optimize_sheet_display_bounds(current_data, None)
            return {
                'type': 'excel',
                'file_type': 'Excel',
                'is_new_file': True,
                'sheets': optimized_sheets,
                'message': '新增的Excel文件'
            }
        
        # 并行对比工作表数据，并优化显示边界
        diff_sheets = self._parallel_compare_sheets_optimized(current_data, previous_data)
        
        # 检查是否有删除的工作表
        for sheet_name in previous_data.keys():
            if sheet_name not in current_data:
                diff_sheets[sheet_name] = {
                    'status': 'deleted',
                    'rows': []
                }
        
        # 更新性能统计
        processing_time = time.time() - start_time
        self.performance_stats['excel_processing_time'] += processing_time
        
        print(f"Excel diff处理完成，耗时: {processing_time:.2f}秒，工作表数量: {len(diff_sheets)}")
        
        # 详细打印每个工作表的数据结构
        for sheet_name, sheet_data in diff_sheets.items():
            print(f"工作表 '{sheet_name}' 数据结构:")
            print(f"  - status: {sheet_data.get('status', 'unknown')}")
            print(f"  - has_changes: {sheet_data.get('has_changes', False)}")
            print(f"  - headers: {sheet_data.get('headers', [])}")
            if 'rows' in sheet_data:
                print(f"  - rows数量: {len(sheet_data['rows'])}")
                if sheet_data['rows']:
                    first_row = sheet_data['rows'][0]
                    print(f"  - 第一行示例: {first_row}")
                    if 'cells' in first_row:
                        print(f"  - 第一行cells数量: {len(first_row['cells'])}")
                        if first_row['cells']:
                            print(f"  - 第一个cell示例: {first_row['cells'][0]}")
                else:
                    print(f"  - rows为空")
        
        result = {
            'type': 'excel',  # 添加type字段供模板识别
            'file_type': 'Excel',
            'file_path': file_path,
            'sheets': diff_sheets,
            'has_changes': any(sheet.get('has_changes', False) for sheet in diff_sheets.values())
        }
        
        print(f"最终返回的Excel diff数据结构: type={result['type']}, sheets数量={len(result['sheets'])}, has_changes={result['has_changes']}")
        
        return result
    
    def _parallel_compare_sheets(self, current_data, previous_data):
        """并行对比多个工作表"""
        diff_sheets = {}
        
        # 准备任务列表
        sheet_tasks = []
        for sheet_name in current_data.keys():
            current_sheet = current_data[sheet_name]
            previous_sheet = previous_data.get(sheet_name, [])
            sheet_tasks.append((sheet_name, current_sheet, previous_sheet))
        
        # 计算总数据量来智能选择处理方式
        total_rows = sum(max(len(current_sheet), len(previous_sheet)) 
                        for _, current_sheet, previous_sheet in sheet_tasks)
        
        # 大幅提高并行处理阈值，避免小数据集的线程开销
        use_parallel = len(sheet_tasks) >= 5 and total_rows >= 10000
        
        if not use_parallel:
            print(f"使用高性能串行处理 {len(sheet_tasks)} 个工作表 (总计 {total_rows} 行)")
            for sheet_name, current_sheet, previous_sheet in sheet_tasks:
                diff_sheets[sheet_name] = self._compare_sheet_data(current_sheet, previous_sheet)
            return diff_sheets
        
        # 多线程并行处理
        print(f"开始并行处理 {len(sheet_tasks)} 个工作表...")
        self.performance_stats['parallel_tasks_count'] += len(sheet_tasks)
        
        executor = self._get_thread_pool()
        # 提交任务
        future_to_sheet = {
            executor.submit(self._compare_sheet_data_safe, current_sheet, previous_sheet): sheet_name
            for sheet_name, current_sheet, previous_sheet in sheet_tasks
        }
        
        # 收集结果
        for future in as_completed(future_to_sheet):
            sheet_name = future_to_sheet[future]
            try:
                result = future.result(timeout=5)  # 减少超时时间到5秒
                diff_sheets[sheet_name] = result
            except Exception as e:
                print(f"工作表 '{sheet_name}' 处理失败: {str(e)}")
                # 失败时使用串行处理作为备选
                current_sheet = next(cs for sn, cs, _ in sheet_tasks if sn == sheet_name)
                previous_sheet = next(ps for sn, _, ps in sheet_tasks if sn == sheet_name)
                diff_sheets[sheet_name] = self._compare_sheet_data(current_sheet, previous_sheet)
        
        return diff_sheets
    
    def _compare_sheet_data_safe(self, current_sheet, previous_sheet):
        """线程安全的工作表对比方法"""
        try:
            return self._compare_sheet_data(current_sheet, previous_sheet)
        except Exception as e:
            print(f"工作表对比出错: {str(e)}")
            # 返回基本的错误结果
            return {
                'status': 'error',
                'headers': [],
                'rows': [],
                'has_changes': False,
                'error_message': str(e)
            }
    
    def _format_excel_sheets_for_display(self, sheets_data):
        """格式化Excel工作表数据用于显示"""
        formatted_sheets = {}
        
        for sheet_name, rows in sheets_data.items():
            if not rows:
                continue
                
            # 获取列名（第一行作为表头）
            headers = list(rows[0].keys()) if rows else []
            
            formatted_sheets[sheet_name] = {
                'headers': headers,
                'rows': rows,
                'status': 'new'
            }
        
        return formatted_sheets
    
    def _optimize_sheet_display_bounds(self, current_data, previous_data):
        """优化工作表显示边界，隐藏重叠的空白区域"""
        optimized_sheets = {}
        
        for sheet_name, current_sheet in current_data.items():
            previous_sheet = previous_data.get(sheet_name, []) if previous_data else []
            
            # 分析当前版本和前一版本的数据边界
            current_bounds = self._analyze_sheet_bounds(current_sheet)
            previous_bounds = self._analyze_sheet_bounds(previous_sheet) if previous_sheet else {'max_row': 0, 'max_col': 0}
            
            # 计算需要显示的最大边界（取两个版本的最大值）
            display_max_row = max(current_bounds['max_row'], previous_bounds['max_row'])
            display_max_col = max(current_bounds['max_col'], previous_bounds['max_col'])
            
            # 裁剪数据到实际需要的边界
            optimized_sheet = self._trim_sheet_to_bounds(current_sheet, display_max_row, display_max_col)
            
            optimized_sheets[sheet_name] = {
                'headers': list(optimized_sheet[0].keys()) if optimized_sheet else [],
                'rows': optimized_sheet,
                'status': 'optimized',
                'bounds': {
                    'max_row': display_max_row,
                    'max_col': display_max_col,
                    'current_bounds': current_bounds,
                    'previous_bounds': previous_bounds
                }
            }
            
            print(f"工作表 {sheet_name} 优化边界: 显示到第{display_max_row}行, 第{display_max_col}列")
        
        return optimized_sheets
    
    def _analyze_sheet_bounds(self, sheet_data):
        """分析工作表的实际数据边界"""
        if not sheet_data:
            return {'max_row': 0, 'max_col': 0}
        
        max_row = 0
        max_col = 0
        
        for row_idx, row_data in enumerate(sheet_data):
            row_has_data = False
            if isinstance(row_data, dict):
                for col_name, cell_value in row_data.items():
                    if cell_value and str(cell_value).strip():
                        row_has_data = True
                        # 计算列索引
                        col_idx = self._column_letter_to_index(col_name)
                        max_col = max(max_col, col_idx)
            
            if row_has_data:
                max_row = row_idx + 1
        
        return {'max_row': max_row, 'max_col': max_col}
    
    def _column_letter_to_index(self, col_letter):
        """将Excel列字母转换为索引 (A->1, B->2, ...)"""
        result = 0
        for char in col_letter:
            result = result * 26 + (ord(char) - ord('A') + 1)
        return result
    
    def _trim_sheet_to_bounds(self, sheet_data, max_row, max_col):
        """裁剪工作表数据到指定边界"""
        if not sheet_data:
            return []
        
        # 生成需要保留的列名列表
        keep_columns = []
        for col_idx in range(1, max_col + 1):
            col_letter = self._get_column_letter(col_idx)
            keep_columns.append(col_letter)
        
        # 裁剪行和列
        trimmed_sheet = []
        for row_idx in range(min(len(sheet_data), max_row)):
            row_data = sheet_data[row_idx]
            if isinstance(row_data, dict):
                trimmed_row = {col: row_data.get(col, '') for col in keep_columns}
                trimmed_sheet.append(trimmed_row)
        
        return trimmed_sheet
    
    def _parallel_compare_sheets_optimized(self, current_data, previous_data):
        """优化的并行工作表对比，包含空白区域处理"""
        diff_sheets = {}
        
        for sheet_name in current_data.keys():
            current_sheet = current_data[sheet_name]
            previous_sheet = previous_data.get(sheet_name, [])
            
            # 分析两个版本的数据边界
            current_bounds = self._analyze_sheet_bounds(current_sheet)
            previous_bounds = self._analyze_sheet_bounds(previous_sheet)
            
            # 计算显示边界
            display_max_row = max(current_bounds['max_row'], previous_bounds['max_row'])
            display_max_col = max(current_bounds['max_col'], previous_bounds['max_col'])
            
            # 裁剪到显示边界
            trimmed_current = self._trim_sheet_to_bounds(current_sheet, display_max_row, display_max_col)
            trimmed_previous = self._trim_sheet_to_bounds(previous_sheet, display_max_row, display_max_col)
            
            # 执行对比
            diff_result = self._compare_sheet_data(trimmed_current, trimmed_previous)
            diff_result['bounds'] = {
                'display_max_row': display_max_row,
                'display_max_col': display_max_col,
                'current_bounds': current_bounds,
                'previous_bounds': previous_bounds
            }
            
            diff_sheets[sheet_name] = diff_result
            
            print(f"工作表 {sheet_name} 对比完成: 显示边界 {display_max_row}行x{display_max_col}列")
        
        return diff_sheets
    
    def _compare_sheet_data(self, current_sheet, previous_sheet):
        """比较两个工作表的数据 - 高性能版本"""
        if not current_sheet and not previous_sheet:
            return {'status': 'unchanged', 'rows': []}
        
        if not previous_sheet:
            # 新增工作表 - 修复表头获取
            if current_sheet and isinstance(current_sheet[0], dict):
                headers = list(current_sheet[0].keys())
            elif current_sheet and isinstance(current_sheet[0], (list, tuple)):
                headers = [f'列{i+1}' for i in range(len(current_sheet[0]))]
            else:
                headers = []
            
            # 转换为前端期望的格式，包含cells字段
            formatted_rows = []
            for i, row in enumerate(current_sheet):
                cells = []
                if isinstance(row, dict):
                    for header in headers:
                        cell_value = row.get(header, '')
                        cells.append({
                            'value': cell_value,
                            'status': 'added'
                        })
                
                formatted_rows.append({
                    'row_number': i + 1,
                    'status': 'added',
                    'cells': cells
                })
            
            return {
                'status': 'new',
                'headers': headers,
                'rows': formatted_rows,
                'has_changes': True
            }
        
        if not current_sheet:
            return {'status': 'deleted', 'rows': []}
        
        # 使用快速比较，确保能检测到变化
        return self._fast_compare_rows(current_sheet, previous_sheet)
    
    @staticmethod
    def _normalize_cell_value(val):
        """标准化单元格值，统一处理NaN/None/空字符串，避免假diff"""
        import math
        if val is None:
            return None
        # 处理 float NaN
        if isinstance(val, float):
            if math.isnan(val):
                return None
        # 尝试 pandas isna
        try:
            import pandas as pd
            if pd.isna(val):
                return None
        except (TypeError, ValueError, ImportError):
            pass
        val_str = str(val).strip()
        if val_str.lower() in ('', 'nan', 'none', 'null', '<na>', 'undefined'):
            return None
        return val_str

    def _fast_compare_rows(self, current_sheet, previous_sheet):
        """快速行比较 - 生成前端兼容的数据格式"""
        # 检测数据格式并获取合并后的表头（处理列删除/新增）
        current_headers = []
        previous_headers = []
        
        if current_sheet and isinstance(current_sheet[0], dict):
            current_headers = list(current_sheet[0].keys())
        elif current_sheet and isinstance(current_sheet[0], (list, tuple)):
            current_headers = [f'列{i+1}' for i in range(len(current_sheet[0]))]
            
        if previous_sheet and isinstance(previous_sheet[0], dict):
            previous_headers = list(previous_sheet[0].keys())
        elif previous_sheet and isinstance(previous_sheet[0], (list, tuple)):
            previous_headers = [f'列{i+1}' for i in range(len(previous_sheet[0]))]
        
        # 合并所有表头，保持顺序
        all_headers = []
        # 先添加前一版本的表头（保持原有顺序）
        for header in previous_headers:
            if header not in all_headers:
                all_headers.append(header)
        # 再添加当前版本新增的表头
        for header in current_headers:
            if header not in all_headers:
                all_headers.append(header)
        
        headers = all_headers
        
        max_rows = max(len(current_sheet), len(previous_sheet))
        diff_rows = []
        has_changes = False
        
        # 值规范化函数，避免NaN/None/空字符串产生假diff
        normalize = self._normalize_cell_value
        
        # 优化的逐行比较，生成单元格级别的变化数据
        for i in range(max_rows):
            if i < len(current_sheet) and i < len(previous_sheet):
                current_row = current_sheet[i]
                previous_row = previous_sheet[i]
                
                # 检查单元格级别的变化（智能列对齐）
                cell_changes = {}
                row_has_changes = False
                
                if isinstance(current_row, dict) and isinstance(previous_row, dict):
                    for header in headers:
                        current_value = current_row.get(header, '')
                        previous_value = previous_row.get(header, '')
                        
                        # 规范化值后再比较，避免NaN vs ''等假diff
                        norm_current = normalize(current_value)
                        norm_previous = normalize(previous_value)
                        
                        # 如果列在当前版本中不存在，标记为删除
                        if header not in current_headers and header in previous_headers:
                            # 仅当旧值非空时才标记为删除
                            if norm_previous is not None:
                                cell_changes[header] = {
                                    'old': str(previous_value),
                                    'new': '',
                                    'status': 'removed'
                                }
                                row_has_changes = True
                        # 如果列在前一版本中不存在，标记为新增
                        elif header in current_headers and header not in previous_headers:
                            # 仅当新值非空时才标记为新增
                            if norm_current is not None:
                                cell_changes[header] = {
                                    'old': '',
                                    'new': str(current_value),
                                    'status': 'added'
                                }
                                row_has_changes = True
                        # 如果规范化后的值不同，才标记为修改
                        elif norm_current != norm_previous:
                            cell_changes[header] = {
                                'old': str(previous_value) if norm_previous is not None else '',
                                'new': str(current_value) if norm_current is not None else '',
                                'status': 'modified'
                            }
                            row_has_changes = True
                
                # 只有当行真正有变化时才添加
                if row_has_changes:
                    
                    diff_rows.append({
                        'row_number': i + 1,
                        'status': 'modified',
                        'row_data': current_row,
                        'previous_data': previous_row,
                        'cell_changes': cell_changes
                    })
                    has_changes = True
                    
            elif i < len(current_sheet):
                # 新增行
                diff_rows.append({
                    'row_number': i + 1,
                    'status': 'added',
                    'row_data': current_sheet[i],
                    'previous_data': {},
                    'cell_changes': {}
                })
                has_changes = True
            else:
                # 删除行
                diff_rows.append({
                    'row_number': i + 1,
                    'status': 'removed',
                    'row_data': {},
                    'previous_data': previous_sheet[i],
                    'cell_changes': {}
                })
                has_changes = True
        
        # 转换为前端期望的格式，包含cells字段
        formatted_rows = []
        for row in diff_rows:
            # 将row_data转换为cells格式
            cells = []
            row_data = row.get('row_data', {})
            if isinstance(row_data, dict):
                for header in headers:
                    cell_value = row_data.get(header, '')
                    cell_changes = row.get('cell_changes', {})
                    if header in cell_changes:
                        cells.append({
                            'value': cell_value,
                            'old_value': cell_changes[header]['old'],
                            'new_value': cell_changes[header]['new'],
                            'status': 'changed'
                        })
                    else:
                        cells.append({
                            'value': cell_value,
                            'status': row['status']
                        })
            
            formatted_row = {
                'row_number': row['row_number'],
                'status': row['status'],
                'cells': cells
            }
            formatted_rows.append(formatted_row)
        
        return {
            'status': 'modified' if has_changes else 'unchanged',
            'headers': headers,
            'rows': formatted_rows,
            'has_changes': has_changes
        }
    
    def _ultra_fast_compare_rows(self, current_sheet, previous_sheet):
        """超快速行比较 - 最小化处理开销"""
        # 获取表头
        if current_sheet and isinstance(current_sheet[0], dict):
            headers = list(current_sheet[0].keys())
        else:
            headers = []
        
        # 快速检查是否有变化 - 更宽松的检查
        has_changes = False
        if len(current_sheet) != len(previous_sheet):
            has_changes = True
        else:
            # 检查前几行是否有变化，避免完整比较
            check_rows = min(10, len(current_sheet), len(previous_sheet))
            for i in range(check_rows):
                if current_sheet[i] != previous_sheet[i]:
                    has_changes = True
                    break
        
        if not has_changes:
            return {
                'status': 'unchanged',
                'headers': headers,
                'rows': [],
                'has_changes': False
            }
        
        # 只处理有变化的行，简化数据结构
        diff_rows = []
        max_rows = max(len(current_sheet), len(previous_sheet))
        
        for i in range(min(50, max_rows)):  # 限制处理前50行以提高性能
            if i < len(current_sheet) and i < len(previous_sheet):
                if current_sheet[i] != previous_sheet[i]:
                    # 简化的单元格变化检测
                    cell_changes = {}
                    if isinstance(current_sheet[i], dict) and isinstance(previous_sheet[i], dict):
                        for key in headers[:10]:  # 只检查前10列
                            curr_val = str(current_sheet[i].get(key, ''))
                            prev_val = str(previous_sheet[i].get(key, ''))
                            if curr_val != prev_val:
                                cell_changes[key] = {'old': prev_val, 'new': curr_val}
                    
                    diff_rows.append({
                        'row_number': i + 1,
                        'status': 'modified',
                        'row_data': current_sheet[i],
                        'previous_data': previous_sheet[i],
                        'cell_changes': cell_changes
                    })
            elif i < len(current_sheet):
                diff_rows.append({
                    'row_number': i + 1,
                    'status': 'added',
                    'row_data': current_sheet[i],
                    'previous_data': {},
                    'cell_changes': {}
                })
            else:
                diff_rows.append({
                    'row_number': i + 1,
                    'status': 'removed',
                    'row_data': {},
                    'previous_data': previous_sheet[i],
                    'cell_changes': {}
                })
        
        return {
            'status': 'modified',
            'headers': headers,
            'rows': diff_rows,
            'has_changes': True
        }
    
    def _parallel_compare_rows(self, current_sheet, previous_sheet, headers, max_rows):
        """并行比较行数据"""
        print(f"使用并行处理比较 {max_rows} 行数据...")
        
        # 分批处理，每批500行
        batch_size = 500
        batches = []
        
        for start_idx in range(0, max_rows, batch_size):
            end_idx = min(start_idx + batch_size, max_rows)
            batches.append((start_idx, end_idx))
        
        diff_rows = []
        has_changes = False
        
        executor = self._get_thread_pool()
        # 提交批处理任务
        future_to_batch = {
            executor.submit(self._compare_row_batch, current_sheet, previous_sheet, headers, start_idx, end_idx): (start_idx, end_idx)
            for start_idx, end_idx in batches
        }
        
        # 收集结果并按顺序合并
        batch_results = {}
        for future in as_completed(future_to_batch):
            start_idx, end_idx = future_to_batch[future]
            try:
                batch_diff_rows, batch_has_changes = future.result(timeout=60)
                batch_results[start_idx] = (batch_diff_rows, batch_has_changes)
                if batch_has_changes:
                    has_changes = True
            except Exception as e:
                print(f"批处理 {start_idx}-{end_idx} 失败: {str(e)}")
                # 失败时使用串行处理
                batch_diff_rows, batch_has_changes = self._compare_row_batch(
                    current_sheet, previous_sheet, headers, start_idx, end_idx
                )
                batch_results[start_idx] = (batch_diff_rows, batch_has_changes)
                if batch_has_changes:
                    has_changes = True
        
        # 按顺序合并结果
        for start_idx in sorted(batch_results.keys()):
            batch_diff_rows, _ = batch_results[start_idx]
            diff_rows.extend(batch_diff_rows)
        
        return {
            'status': 'modified' if has_changes else 'unchanged',
            'headers': headers,
            'rows': diff_rows,
            'has_changes': has_changes
        }
    
    def _compare_row_batch(self, current_sheet, previous_sheet, headers, start_idx, end_idx):
        """比较一批行数据"""
        diff_rows = []
        has_changes = False
        
        for i in range(start_idx, end_idx):
            current_row = current_sheet[i] if i < len(current_sheet) else {}
            previous_row = previous_sheet[i] if i < len(previous_sheet) else {}
            
            if not current_row and previous_row:
                # 删除的行
                diff_rows.append({
                    'row_number': i + 1,
                    'status': 'removed',
                    'previous_data': previous_row
                })
                has_changes = True
            elif current_row and not previous_row:
                # 新增的行
                diff_rows.append({
                    'row_number': i + 1,
                    'status': 'added',
                    'row_data': current_row
                })
                has_changes = True
            elif current_row != previous_row:
                # 修改的行
                cell_changes = {}
                for col in headers:
                    current_val = current_row.get(col, '')
                    previous_val = previous_row.get(col, '')
                    if current_val != previous_val:
                        cell_changes[col] = {
                            'old': previous_val,
                            'new': current_val
                        }
                
                if cell_changes:
                    diff_rows.append({
                        'row_number': i + 1,
                        'status': 'modified',
                        'row_data': current_row,
                        'previous_data': previous_row,
                        'cell_changes': cell_changes
                    })
                    has_changes = True
                else:
                    # 无变化的行
                    diff_rows.append({
                        'row_number': i + 1,
                        'status': 'unchanged',
                        'row_data': current_row
                    })
            else:
                # 无变化的行
                diff_rows.append({
                    'row_number': i + 1,
                    'status': 'unchanged',
                    'row_data': current_row
                })
        
        return diff_rows, has_changes
    
    def _serial_compare_rows(self, current_sheet, previous_sheet, headers, max_rows):
        """串行比较行数据（用于小数据集）"""
        diff_rows = []
        has_changes = False
        
        for i in range(max_rows):
            current_row = current_sheet[i] if i < len(current_sheet) else {}
            previous_row = previous_sheet[i] if i < len(previous_sheet) else {}
            
            if not current_row and previous_row:
                # 删除的行
                diff_rows.append({
                    'row_number': i + 1,
                    'status': 'removed',
                    'previous_data': previous_row
                })
                has_changes = True
            elif current_row and not previous_row:
                # 新增的行
                diff_rows.append({
                    'row_number': i + 1,
                    'status': 'added',
                    'row_data': current_row
                })
                has_changes = True
            elif current_row != previous_row:
                # 修改的行
                cell_changes = {}
                for col in headers:
                    current_val = current_row.get(col, '')
                    previous_val = previous_row.get(col, '')
                    if current_val != previous_val:
                        cell_changes[col] = {
                            'old': previous_val,
                            'new': current_val
                        }
                
                if cell_changes:
                    diff_rows.append({
                        'row_number': i + 1,
                        'status': 'modified',
                        'row_data': current_row,
                        'previous_data': previous_row,
                        'cell_changes': cell_changes
                    })
                    has_changes = True
                else:
                    # 无变化的行
                    diff_rows.append({
                        'row_number': i + 1,
                        'status': 'unchanged',
                        'row_data': current_row
                    })
            else:
                # 无变化的行
                diff_rows.append({
                    'row_number': i + 1,
                    'status': 'unchanged',
                    'row_data': current_row
                })
        
        return {
            'status': 'modified' if has_changes else 'unchanged',
            'headers': headers,
            'rows': diff_rows,
            'has_changes': has_changes
        }
    
    def get_performance_stats(self):
        """获取性能统计信息"""
        return {
            'total_diff_time': f"{self.performance_stats['total_diff_time']:.2f}秒",
            'excel_processing_time': f"{self.performance_stats['excel_processing_time']:.2f}秒",
            'git_operations_time': f"{self.performance_stats['git_operations_time']:.2f}秒",
            'parallel_tasks_count': self.performance_stats['parallel_tasks_count'],
            'max_workers': self.max_workers,
            'cpu_count': os.cpu_count()
        }
    
    def cleanup_thread_pool(self):
        """清理线程池资源"""
        try:
            if hasattr(self, 'thread_pool') and self.thread_pool:
                self.thread_pool.shutdown(wait=True)
                print("🔧 [GIT_SERVICE] 线程池已清理")
                self.thread_pool = None
        except Exception as e:
            print(f"🔧 [GIT_SERVICE] 清理线程池时出错: {str(e)}")
    
    def __del__(self):
        """析构函数，确保资源清理"""
        self.cleanup_thread_pool()
    
    def _parse_unified_diff(self, patch_text):
        """解析unified diff格式，返回结构化的hunks数据"""
        hunks = []
        lines = patch_text.split('\n')
        current_hunk = None
        
        for line in lines:
            # 匹配hunk头部，如 @@ -31,17 +36,20 @@
            if line.startswith('@@'):
                if current_hunk:
                    hunks.append(current_hunk)
                
                # 解析hunk头部信息
                import re
                match = re.match(r'@@ -(\d+),?(\d*) \+(\d+),?(\d*) @@(.*)', line)
                if match:
                    old_start = int(match.group(1))
                    old_count = int(match.group(2)) if match.group(2) else 1
                    new_start = int(match.group(3))
                    new_count = int(match.group(4)) if match.group(4) else 1
                    context = match.group(5).strip()
                    
                    current_hunk = {
                        'header': line,
                        'old_start': old_start,
                        'old_count': old_count,
                        'new_start': new_start,
                        'new_count': new_count,
                        'context': context,
                        'lines': []
                    }
            elif current_hunk and (line.startswith(' ') or line.startswith('+') or line.startswith('-')):
                # 解析diff行
                if line.startswith(' '):
                    line_type = 'context'
                elif line.startswith('+'):
                    line_type = 'added'
                elif line.startswith('-'):
                    line_type = 'removed'
                else:
                    continue
                
                current_hunk['lines'].append({
                    'type': line_type,
                    'content': line[1:] if len(line) > 0 else '',  # 去掉前缀符号
                    'raw': line
                })
                # print(f"解析diff行: {line_type} - {line[:50]}")
        
        # 添加最后一个hunk
        if current_hunk:
            hunks.append(current_hunk)
        
        return hunks

    def _compare_dataframes(self, old_df, new_df, sheet_name):
        """比较两个DataFrame的差异"""
        changes = []
        
        try:
            # 转换为字符串以便比较
            old_df = old_df.astype(str).fillna('')
            new_df = new_df.astype(str).fillna('')
            
            # 比较行数变化
            old_rows, old_cols = old_df.shape
            new_rows, new_cols = new_df.shape
            
            if new_rows > old_rows:
                for i in range(old_rows, new_rows):
                    row_data = {}
                    for j, col in enumerate(new_df.columns):
                        if j < len(new_df.columns):
                            row_data[chr(65 + j)] = new_df.iloc[i, j] if i < len(new_df) else ''
                    
                    changes.append({
                        'type': 'added',
                        'sheet_name': sheet_name,
                        'row': i + 1,
                        'data': row_data,
                        'message': f'{sheet_name} 第{i+1}行新增'
                    })
            
                
                min_rows = min(old_rows, new_rows)
                for i in range(min_rows):
                    row_changed = False
                    old_row_data = {}
                    new_row_data = {}
                    
        except Exception as e:
            print(f"DataFrame比较失败: {str(e)}")
            return []
        
        return changes
    
    def _generate_basic_diff(self, previous_content, current_content, file_path):
        """生成基本的diff结构"""
        try:
            import difflib
            
            previous_lines = previous_content.splitlines() if previous_content else []
            current_lines = current_content.splitlines() if current_content else []
            
            # 使用difflib生成unified diff
            diff_lines = list(difflib.unified_diff(
                previous_lines, 
                current_lines, 
                fromfile=f'a/{file_path}', 
                tofile=f'b/{file_path}',
                lineterm=''
            ))
            
            if not diff_lines:
                return None
                
            patch_text = '\n'.join(diff_lines)
            hunks = self._parse_unified_diff(patch_text)
            
            return {
                'type': 'code',
                'file_path': file_path,
                'patch': patch_text,
                'hunks': hunks
            }
            
        except Exception as e:
            print(f"生成基本diff失败: {str(e)}")
            return None
    
    def _generate_initial_commit_diff(self, current_content, file_path):
        """生成初始提交的diff（所有内容都是新增的）"""
        try:
            lines = current_content.splitlines() if current_content else []
            
            # 创建一个表示全新文件的hunk
            hunk = {
                'header': f'@@ -0,0 +1,{len(lines)} @@',
                'old_start': 0,
                'old_count': 0,
                'new_start': 1,
                'new_count': len(lines),
                'context': '',
                'lines': []
            }
            
            # 所有行都标记为新增
            for i, line in enumerate(lines):
                hunk['lines'].append({
                    'type': 'added',
                    'content': line,
                    'raw': f'+{line}',
                    'old_line_number': None,
                    'new_line_number': i + 1
                })
            
            # 生成patch文本
            patch_lines = [f'--- /dev/null', f'+++ b/{file_path}', hunk['header']]
            for line_info in hunk['lines']:
                patch_lines.append(line_info['raw'])
            
            patch_text = '\n'.join(patch_lines)
            
            return {
                'type': 'code',
                'file_path': file_path,
                'patch': patch_text,
                'hunks': [hunk]
            }
            
        except Exception as e:
            print(f"生成初始提交diff失败: {str(e)}")
            return None

"""Excel parsing helpers extracted from GitService."""

from __future__ import annotations

import os
import time

import git

GIT_EXCEL_PARSE_ERRORS = (RuntimeError, ValueError, TypeError, AttributeError, KeyError, OSError)
GIT_EXCEL_REPO_INIT_ERRORS = (
    git.exc.GitError,
    OSError,
    RuntimeError,
    ValueError,
    TypeError,
    AttributeError,
)
GIT_EXCEL_WORKBOOK_PARSE_ERRORS = (
    ImportError,
    RuntimeError,
    ValueError,
    TypeError,
    AttributeError,
    KeyError,
    OSError,
)
GIT_EXCEL_EXTRACT_ERRORS = (
    RuntimeError,
    ValueError,
    TypeError,
    AttributeError,
    KeyError,
    OSError,
)
GIT_EXCEL_SIMPLE_EXTRACT_ERRORS = (RuntimeError, ValueError, TypeError, OSError)


def parse_excel_diff(service, commit_id, file_path):
    """Parse Excel diff for a file at a commit."""
    try:
        if not os.path.exists(service.local_path):
            return None

        repo = git.Repo(service.local_path)
        try:
            commit = repo.commit(commit_id)
            try:
                commit.tree[file_path]
                file_exists_in_current = True
            except KeyError:
                file_exists_in_current = False

            if not file_exists_in_current:
                return {
                    "type": "excel",
                    "file_type": "Excel",
                    "operation": "deleted",
                    "message": "该Excel文件已被删除",
                    "file_path": file_path,
                }

            current_data = service._extract_excel_data(commit, file_path)
            previous_data = None
            if commit.parents:
                parent_commit = commit.parents[0]
                try:
                    previous_data = service._extract_excel_data(parent_commit, file_path)
                except KeyError:
                    pass
            return service._generate_excel_diff_data(current_data, previous_data, file_path)
        except GIT_EXCEL_PARSE_ERRORS as exc:
            return {
                "type": "excel",
                "file_type": "Excel",
                "error": True,
                "message": f"无法解析Excel差异: {str(exc)}",
            }
    except GIT_EXCEL_REPO_INIT_ERRORS:
        return None


def extract_excel_data(service, commit, file_path):
    """Extract workbook sheet data from a commit blob."""
    try:
        try:
            blob = commit.tree / file_path
        except KeyError:
            return None

        excel_data = blob.data_stream.read()
        import io
        import warnings

        warnings.filterwarnings(
            "ignore",
            message="Data Validation extension is not supported and will be removed",
        )
        from openpyxl import load_workbook

        excel_file = io.BytesIO(excel_data)
        sheets_data = {}
        try:
            workbook = load_workbook(excel_file, data_only=True)
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                sheet_rows = []
                max_row = worksheet.max_row
                max_col = worksheet.max_column

                if max_row > 0 and max_col > 0:
                    has_content = False
                    for row in range(1, min(max_row + 1, 11)):
                        for col in range(1, min(max_col + 1, 11)):
                            cell_value = worksheet.cell(row=row, column=col).value
                            if cell_value is not None and str(cell_value).strip():
                                has_content = True
                                break
                        if has_content:
                            break

                    if has_content:
                        actual_bounds = service._detect_data_bounds(worksheet, max_row, max_col)
                        actual_max_row = actual_bounds["max_row"]
                        actual_max_col = actual_bounds["max_col"]
                        for row in range(1, actual_max_row + 1):
                            row_data = {}
                            for col in range(1, actual_max_col + 1):
                                cell_value = worksheet.cell(row=row, column=col).value
                                col_letter = service._get_column_letter(col)
                                cell_str = str(cell_value) if cell_value is not None else ""
                                row_data[col_letter] = cell_str
                            sheet_rows.append(row_data)
                sheets_data[sheet_name] = sheet_rows
            workbook.close()
            return sheets_data
        except GIT_EXCEL_WORKBOOK_PARSE_ERRORS:
            return service._extract_excel_data_simple(excel_data, file_path)
    except GIT_EXCEL_EXTRACT_ERRORS:
        return None


def get_column_letter(col_num):
    """Convert numeric index to Excel column letter."""
    result = ""
    while col_num > 0:
        col_num -= 1
        result = chr(col_num % 26 + ord("A")) + result
        col_num //= 26
    return result


def detect_data_bounds(worksheet, max_row, max_col):
    """Detect real data bounds for worksheet."""
    actual_max_row = 0
    actual_max_col = 0

    for row in range(max_row, 0, -1):
        has_data = False
        for col in range(1, max_col + 1):
            cell_value = worksheet.cell(row=row, column=col).value
            if cell_value is not None and str(cell_value).strip():
                has_data = True
                break
        if has_data:
            actual_max_row = row
            break

    for col in range(max_col, 0, -1):
        has_data = False
        for row in range(1, actual_max_row + 1):
            cell_value = worksheet.cell(row=row, column=col).value
            if cell_value is not None and str(cell_value).strip():
                has_data = True
                break
        if has_data:
            actual_max_col = col
            break

    actual_max_col = max(actual_max_col, 10)
    actual_max_row = max(actual_max_row, 1)
    return {"max_row": actual_max_row, "max_col": actual_max_col}


def extract_excel_data_simple(excel_data, file_path):
    """Best-effort fallback when workbook parsing fails."""
    try:
        return {
            "Sheet1": [
                {
                    "A": f"Excel文件: {file_path}",
                    "B": f"文件大小: {len(excel_data)} 字节",
                    "C": "由于兼容性问题，无法显示详细内容",
                }
            ]
        }
    except GIT_EXCEL_SIMPLE_EXTRACT_ERRORS:
        return None


def generate_excel_diff_data(service, current_data, previous_data, file_path):
    """Generate Excel diff payload for UI rendering."""
    start_time = time.time()

    if not current_data:
        return {
            "type": "excel",
            "file_type": "Excel",
            "error": True,
            "message": "无法读取Excel文件内容",
        }

    if not previous_data:
        optimized_sheets = service._optimize_sheet_display_bounds(current_data, None)
        return {
            "type": "excel",
            "file_type": "Excel",
            "is_new_file": True,
            "sheets": optimized_sheets,
            "message": "新增的Excel文件",
        }

    diff_sheets = service._parallel_compare_sheets_optimized(current_data, previous_data)
    for sheet_name in previous_data.keys():
        if sheet_name not in current_data:
            diff_sheets[sheet_name] = {"status": "deleted", "rows": []}

    processing_time = time.time() - start_time
    service.performance_stats["excel_processing_time"] += processing_time

    return {
        "type": "excel",
        "file_type": "Excel",
        "file_path": file_path,
        "sheets": diff_sheets,
        "has_changes": any(sheet.get("has_changes", False) for sheet in diff_sheets.values()),
    }
